from pyzabbix import ZabbixAPI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import time, os, getpass

# VARIAVEIS
ZBX_IP = input("IP Zabbix: ")
USER = input("Login: ")
PASS = getpass.getpass("Senha: ")

if ZBX_IP == "":
    print ("Favor inserir o endereço IP")
    exit()

if USER == "":
    print ("Favor inserir o Login")
    exit()


# CONECTANDO AO ZABBIX 
zapi = ZabbixAPI('http://'+ZBX_IP+'/zabbix')
zapi.login(USER,PASS)
print("--> Conectado com sucesso!")

# CONFIGURANDO PERIODO DE EXTRACAO DOS DADOS
DATA_INI = input("Inserir data de inicio (DD/MM/AAAA): ")
DATA_FINAL = input("Inserir data de inicio (DD/MM/AAAA): ")

DATA_INICIO = time.mktime(time.strptime('{INICIO} 00:00:00'.format(INICIO=DATA_INI),'%d/%m/%Y %H:%M:%S'))
DATA_FIM = time.mktime(time.strptime('{FIM} 23:59:59'.format(FIM=DATA_FINAL),'%d/%m/%Y %H:%M:%S'))

# DEFINIDO A SEVERIDADE MINIMA
print("""
Qual a severidade minima dos eventos:

0 - Not classified
1 - Information
2 - Warning
3 - Average
4 - High
5 - Disaster

""")
SEV_MIN = input("Insira o numero: ")

# CRIANDO PLANILHA
wb = Workbook()
sheet = wb.active
sheet.title = "Eventos"
sheet['A1'] = "DATA"
sheet['A1'].font = Font(sz=12, bold=True)
sheet['B1'] = "HOST"
sheet['B1'].font = Font(sz=12, bold=True)
sheet['C1'] = "TRIGGER"
sheet['C1'].font = Font(sz=12, bold=True)
sheet['D1'] = "SEVERIDADE"
sheet['D1'].font = Font(sz=12, bold=True)
sheet['E1'] = "STATUS"
sheet['E1'].font = Font(sz=12, bold=True)
sheet['F1'] = "GRUPO"
sheet['F1'].font = Font(sz=12, bold=True)
sheet['G1'] = "ACK"
sheet['G1'].font = Font(sz=12, bold=True)




# VALOR EVENTO
EVENT = {
    "0": "OK",
    "1": "PROBLEM"
}

# SEVERIDADE
SEV = {
    "0": "Not classified",
    "1": "Information",
    "2": "Warning",
    "3": "Average",
    "4": "High",
    "5": "Disaster"
}

# ACKNOWLEDGED
ACK = {
    "0": "Não",
    "1": "Sim"
}

row=2
eventos=0
inicio=time.strftime("%d/%m/%Y %H:%M:%S")
for e in zapi.event.get(output="extend",time_from=DATA_INICIO,time_till=DATA_FIM,sortfield=['clock'],sortorder="ASC",value=1):
    for t in zapi.trigger.get(output="extend",triggerids=e['objectid'],expandDescription=True,min_severity=int(SEV_MIN)):
        for h in zapi.host.get(output="extend",triggerids=e['objectid'],selectGroups=['name']):
            data = time.strftime("%d/%m/%Y %H:%M:%S", time.localtime(int(e['clock'])))
            os.system('cls' if os.name == 'nt' else 'clear')
            print(time.strftime("%d/%m/%Y %H:%M:%S"), "Eventos: {0}".format(eventos))
            sheet.cell(row=row, column=1).value = data
            sheet.cell(row=row, column=2).value = h['host']
            sheet.cell(row=row, column=3).value = t['description']
            sheet.cell(row=row, column=4).value = SEV[t['priority']]
            sheet.cell(row=row, column=5).value = EVENT[e['value']]
            sheet.cell(row=row, column=6).value = h['groups'][0]['name']
            sheet.cell(row=row, column=7).value = ACK[e['acknowledged']]
            row+=1
            eventos+=1
            
fim=time.strftime("%d/%m/%Y %H:%M:%S")
print("Inicio da execucao: {0}\nFim da execucao: {1}\nEventos coletados: {2}".format(inicio,fim,eventos))
wb.save('EXPORT_EVENT_1.xlsx')
os.system('PAUSE')
zapi.user.logout()